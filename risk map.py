import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LogNorm
from scipy.stats import gaussian_kde, pearsonr
import openpyxl
import warnings
import time as _time

warnings.filterwarnings("ignore")

# ==============================
# User Configuration
# ==============================

DATA_ROOT = r"E:\Claude code\WORK-2\Data"
TS_ROOT   = os.path.join(DATA_ROOT, "Short_Dose", "TIME_SERIES")
EXCEL_PATH = os.path.join(DATA_ROOT, "source term-SORACA-process.xlsx")
OUTPUT_DIR = DATA_ROOT

DOSE_TIMES = [15, 30, 45]          # dose snapshot times (min from plume start)
EVAC_TIMES = [15, 30, 45]          # evacuation times to evaluate (min from depart)
DEPART_DELAY = 0.5                 # hours (30 min)

GRID_SIZE  = 100
CELL_SIZE_M = 400.0
PLANT_X    = 247413.0              # UTM Easting (m)
PLANT_Y    = 2501099.0             # UTM Northing (m)
UTM_ZONE   = "50N"
CONFIDENCE_LEVEL = 0.99

N_SAMPLES  = 4996                  # must match number of scenario files
RANDOM_SEED = 42

plt.rcParams['font.family'] = 'Times New Roman'

# ==============================
# Step 1: Read Source-Term Timing
# ==============================

print("=" * 65)
print("  CVaR Risk Map with Dual Uncertainty")
print("  (Onset-of-Release x GE-Declaration timing adjustment)")
print("=" * 65)

print("\n[Step 1] Reading source-term timing data ...")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Sheet1']

categories, onset_vals, ge_vals, rel_freqs = [], [], [], []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    name   = row[0].value
    onset  = row[11].value        # col L: Onset of release (hr)
    ge     = row[12].value        # col M: GE declaration (hr)
    rfreq  = row[2].value         # col C: relative frequency
    if name is None or onset is None or ge is None:
        continue
    categories.append(str(name).strip())
    onset_vals.append(float(onset))
    ge_vals.append(float(ge))
    rel_freqs.append(float(rfreq) if rfreq else 0.0)

onset_arr = np.array(onset_vals)
ge_arr    = np.array(ge_vals)
offset_arr = onset_arr - ge_arr - DEPART_DELAY

print(f"  {len(categories)} source-term categories loaded:")
print(f"  {'Category':12s}  Onset(h)  GE(h)  Offset(h)  Offset(min)")
for c, o, g, off in zip(categories, onset_arr, ge_arr, offset_arr):
    print(f"  {c:12s}  {o:7.1f}   {g:5.2f}   {off:8.2f}    {off*60:8.1f}")

# ==============================
# Step 2: Pearson Correlation
# ==============================

print("\n[Step 2] Pearson correlation ...")
corr, pval = pearsonr(onset_arr, ge_arr)
print(f"  r = {corr:.6f},  p = {pval:.4e}")
strength = "Strong" if abs(corr) > 0.7 else ("Moderate" if abs(corr) > 0.4 else "Weak")
print(f"  => {strength} positive correlation  -->  joint 2D-KDE sampling required")

corr_path = os.path.join(OUTPUT_DIR, "pearson_correlation.txt")
with open(corr_path, "w", encoding="utf-8") as f:
    f.write(f"Pearson Correlation: Onset-of-Release vs GE-Declaration\n")
    f.write(f"N = {len(onset_arr)} categories\n")
    f.write(f"r = {corr:.6f}\n")
    f.write(f"p = {pval:.4e}\n")
    f.write(f"Strength: {strength}\n")
print(f"  Saved: {corr_path}")

# ==============================
# Step 3: 2D KDE  +  Sampling
# ==============================

print("\n[Step 3] Building 2D KDE and sampling ...")

data_2d = np.vstack([onset_arr, ge_arr])
kde = gaussian_kde(data_2d, bw_method='silverman')

np.random.seed(RANDOM_SEED)

# Rejection sampling: Onset > GE > 0
accepted = []
while len(accepted) < N_SAMPLES:
    batch = kde.resample(N_SAMPLES * 3).T
    mask  = (batch[:, 0] > 0) & (batch[:, 1] > 0) & (batch[:, 0] >= batch[:, 1])
    accepted.extend(batch[mask].tolist())

samples = np.array(accepted[:N_SAMPLES])
onset_s   = samples[:, 0]
ge_s      = samples[:, 1]
offset_hr = onset_s - ge_s - DEPART_DELAY
offset_min = offset_hr * 60

print(f"  Sampled {N_SAMPLES} valid (Onset, GE) pairs")
print(f"  Plume-offset range: [{offset_hr.min():.2f}, {offset_hr.max():.2f}] hr")
print(f"                    = [{offset_min.min():.1f}, {offset_min.max():.1f}] min")
n_neg = np.sum(offset_min < 0)
print(f"  Negative offsets (plume before evac): {n_neg} / {N_SAMPLES}")

csv_path = os.path.join(OUTPUT_DIR, "kde_timing_samples.csv")
pd.DataFrame({
    'Scenario_ID':          np.arange(1, N_SAMPLES + 1),
    'Onset_of_Release_hr':  onset_s,
    'GE_Declaration_hr':    ge_s,
    'Depart_Delay_hr':      DEPART_DELAY,
    'Plume_Offset_hr':      offset_hr,
    'Plume_Offset_min':     offset_min,
}).to_csv(csv_path, index=False)
print(f"  Saved: {csv_path}")

# --- KDE scatter plot ---
fig_kde, axes = plt.subplots(1, 2, figsize=(14, 5))

ax = axes[0]
ax.scatter(onset_arr, ge_arr, c='red', s=80, zorder=5, edgecolors='white',
           label='Source-term categories')
ax.scatter(onset_s, ge_s, c='steelblue', s=4, alpha=0.3, label='KDE samples')
ax.set_xlabel('Onset of Release (hr)')
ax.set_ylabel('GE Declaration (hr)')
ax.set_title(f'2D KDE Sampling (r={corr:.3f})')
ax.legend(fontsize=8)

ax = axes[1]
ax.hist(offset_min, bins=60, color='steelblue', edgecolor='white', alpha=0.8)
ax.axvline(0, color='red', ls='--', lw=1.5, label='Evac start (t=0)')
ax.set_xlabel('Plume Offset (min)')
ax.set_ylabel('Frequency')
ax.set_title('Plume Offset Distribution')
ax.legend()

fig_kde.tight_layout()
kde_fig_path = os.path.join(OUTPUT_DIR, "kde_sampling_diagnostic.png")
fig_kde.savefig(kde_fig_path, dpi=200)
plt.close(fig_kde)
print(f"  Saved: {kde_fig_path}")

# ==============================
# Step 4: Load Dose Data
# ==============================

print("\n[Step 4] Loading dose data (3 time-points x 4996 scenarios) ...")

def load_dose_folder(folder, grid_size=100):
    """Load all CSV dose grids from *folder*, return shape (N, grid, grid)."""
    fnames = sorted(
        [f for f in os.listdir(folder) if f.endswith('.csv')],
        key=lambda x: int(os.path.splitext(x)[0])
    )
    n = len(fnames)
    arr = np.zeros((n, grid_size, grid_size), dtype=np.float64)
    t0 = _time.time()
    for idx, fn in enumerate(fnames):
        path = os.path.join(folder, fn)
        with open(path, 'r') as fh:
            lines = fh.readlines()
        for ri, line in enumerate(lines[1:grid_size + 1]):
            fields = line.split(',')
            for ci, v in enumerate(fields[:grid_size]):
                v = v.strip()
                if v:
                    try:
                        arr[idx, ri, ci] = float(v)
                    except (ValueError, OverflowError):
                        pass
        arr[idx] = arr[idx, ::-1, :]          # flip rows
        if (idx + 1) % 1000 == 0:
            elapsed = _time.time() - t0
            print(f"    {idx+1}/{n}  ({elapsed:.1f}s)")
    elapsed = _time.time() - t0
    print(f"    Done: {n} files in {elapsed:.1f}s")
    return arr, fnames

dose = {}
file_lists = {}
for t in DOSE_TIMES:
    folder = os.path.join(TS_ROOT, str(t), "effect")
    print(f"  Loading {t} min ...")
    dose[t], file_lists[t] = load_dose_folder(folder)
    print(f"    shape = {dose[t].shape}")

# Verify consistency
ns = [dose[t].shape[0] for t in DOSE_TIMES]
assert all(n == ns[0] for n in ns), f"Inconsistent scenario counts: {ns}"
N_ACTUAL = ns[0]
assert N_ACTUAL == N_SAMPLES, f"File count {N_ACTUAL} != sample count {N_SAMPLES}"
print(f"  All time-points: {N_ACTUAL} scenarios (consistent)")

# ==============================
# Step 5: Compute CVaR Risk Maps
# ==============================

print("\n[Step 5] Computing CVaR risk maps with timing adjustment ...")

def cvar(doses, cl=0.99):
    """Return (VaR, CVaR) at confidence level *cl*."""
    if len(doses) == 0 or np.all(doses == 0):
        return 0.0, 0.0
    sd = np.sort(doses)
    vi = min(int(np.ceil(cl * len(sd))) - 1, len(sd) - 1)
    var_val = sd[vi]
    tail = sd[sd >= var_val]
    return var_val, (np.mean(tail) if len(tail) else var_val)

d15 = dose[15]   # (N, 100, 100)
d30 = dose[30]
d45 = dose[45]

all_risk = {}
all_var  = {}

for evac_t in EVAC_TIMES:
    t0 = _time.time()
    threshold = 0.05 / 7 / 16 / 24 * (evac_t / 60)
    print(f"\n  --- Evacuation time = {evac_t} min ---")
    print(f"  Risk threshold = {threshold:.6e} Sv")

    # effective plume-exposure time per scenario (minutes)
    eff = evac_t - offset_min                           # (N,)
    n_zero = np.sum(eff <= 0)
    print(f"  Scenarios with zero exposure (plume not yet arrived): "
          f"{n_zero}/{N_SAMPLES} ({100*n_zero/N_SAMPLES:.1f}%)")

    risk_map = np.zeros((GRID_SIZE, GRID_SIZE))
    var_map  = np.zeros((GRID_SIZE, GRID_SIZE))
    risk_cnt = 0

    for i in range(GRID_SIZE):
        for j in range(GRID_SIZE):
            v15 = d15[:, i, j]
            v30 = d30[:, i, j]
            v45 = d45[:, i, j]

            # Piecewise-linear interpolation based on effective exposure time
            adj = np.where(
                eff <= 0,  0.0,
                np.where(eff <= 15, v15 * (eff / 15.0),
                np.where(eff <= 30, v15 + (v30 - v15) * (eff - 15.0) / 15.0,
                np.where(eff <= 45, v30 + (v45 - v30) * (eff - 30.0) / 15.0,
                         v45))))

            vv, cv = cvar(adj, CONFIDENCE_LEVEL)
            var_map[i, j] = vv
            if vv >= threshold:
                risk_map[i, j] = cv
                risk_cnt += 1

    all_risk[evac_t] = risk_map
    all_var[evac_t]  = var_map

    elapsed = _time.time() - t0
    print(f"  Risk cells: {risk_cnt}/{GRID_SIZE**2} ({100*risk_cnt/GRID_SIZE**2:.2f}%)")
    if risk_cnt > 0:
        nz = risk_map[risk_map > 0]
        print(f"  CVaR range: [{nz.min():.4e}, {nz.max():.4e}] Sv")
        print(f"  CVaR mean:  {nz.mean():.4e} Sv")
    print(f"  Elapsed: {elapsed:.1f}s")

# ==============================
# Step 6: Save Results
# ==============================

print("\n[Step 6] Saving results ...")

for evac_t in EVAC_TIMES:
    rm = all_risk[evac_t]
    vm = all_var[evac_t]

    p1 = os.path.join(OUTPUT_DIR, f"cvar_risk_map_{evac_t}.xlsx")
    pd.DataFrame(rm[::-1]).to_excel(p1, index=False, header=False)
    print(f"  {p1}")

    p2 = os.path.join(OUTPUT_DIR, f"cvar_risk_map_{evac_t}_VAR.xlsx")
    pd.DataFrame(vm[::-1]).to_excel(p2, index=False, header=False)
    print(f"  {p2}")

# ==============================
# Step 7: Plot Risk Maps
# ==============================

print("\n[Step 7] Plotting risk maps ...")

try:
    from pyproj import CRS, Transformer
    import contextily as ctx
    HAS_GEO = True
except ImportError:
    HAS_GEO = False
    print("  (contextily/pyproj not installed -- plain plot)")

# Global colorbar range across all time-points
all_nz = np.concatenate([rm[rm > 0].ravel() for rm in all_risk.values()
                         if np.any(rm > 0)])
if len(all_nz):
    g_vmin = max(np.percentile(all_nz, 1),
                 min(0.05/7/16/24*(t/60) for t in EVAC_TIMES))
    g_vmax = np.percentile(all_nz, 99)
else:
    g_vmin, g_vmax = 1e-10, 1e-6

for evac_t in EVAC_TIMES:
    risk_map = all_risk[evac_t]
    threshold = 0.05 / 7 / 16 / 24 * (evac_t / 60)

    half = GRID_SIZE // 2
    x_c = PLANT_X + (np.arange(GRID_SIZE) - half) * CELL_SIZE_M
    y_c = PLANT_Y + (np.arange(GRID_SIZE) - half) * CELL_SIZE_M

    if HAS_GEO:
        crs_utm   = CRS(f"+proj=utm +zone=50 +north +ellps=WGS84 +datum=WGS84 +units=m")
        crs_wgs   = CRS("EPSG:4326")
        tr        = Transformer.from_crs(crs_utm, crs_wgs, always_xy=True)
        lo0, la0  = tr.transform(x_c[0] - CELL_SIZE_M/2, y_c[0] - CELL_SIZE_M/2)
        lo1, la1  = tr.transform(x_c[-1]+ CELL_SIZE_M/2, y_c[-1]+ CELL_SIZE_M/2)
        extent = [lo0, lo1, la0, la1]
    else:
        extent = [x_c[0], x_c[-1], y_c[0], y_c[-1]]

    fig, ax = plt.subplots(figsize=(12, 10), dpi=200)
    ax.set_xlim(extent[0], extent[1])
    ax.set_ylim(extent[2], extent[3])

    if HAS_GEO:
        try:
            ctx.add_basemap(ax, crs=crs_wgs, source=ctx.providers.CartoDB.Positron,
                            attribution=False)
        except Exception:
            pass

    eps = 1e-20
    norm = LogNorm(vmin=max(g_vmin, eps), vmax=max(g_vmax, g_vmin*1.1))
    cmap = plt.get_cmap('rainbow')
    rgba = cmap(norm(np.where(risk_map > 0, risk_map, np.nan)))
    rgba[:, :, 3] = np.where(risk_map > 0, 0.6, 0.0)

    ax.imshow(rgba, extent=extent, origin='lower', interpolation='nearest', zorder=2)

    sm = plt.cm.ScalarMappable(norm=norm, cmap=cmap)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, shrink=0.6)
    cbar.set_label(f'CVaR (Sv)  [{int(CONFIDENCE_LEVEL*100)}% CL]')

    ax.set_title(f'CVaR Risk Map  |  Evac Time = {evac_t} min\n'
                 f'Threshold = {threshold:.4e} Sv  |  Dual-Uncertainty Adjusted',
                 fontsize=13)
    ax.set_xlabel('Longitude' if HAS_GEO else 'UTM Easting (m)')
    ax.set_ylabel('Latitude'  if HAS_GEO else 'UTM Northing (m)')

    fig.tight_layout()
    png = os.path.join(OUTPUT_DIR, f"cvar_risk_map_{evac_t}.png")
    fig.savefig(png, dpi=200, bbox_inches='tight')
    plt.close(fig)
    print(f"  {png}")

# ==============================
# Step 8: Summary
# ==============================

print("\n" + "=" * 65)
print("  SUMMARY")
print("=" * 65)
print(f"  Pearson r(Onset, GE) = {corr:.4f}")
print(f"  Plume offset range   = [{offset_min.min():.1f}, {offset_min.max():.1f}] min")
print(f"  Negative offsets     = {n_neg} / {N_SAMPLES}")
for evac_t in EVAC_TIMES:
    rm = all_risk[evac_t]
    rc = np.sum(rm > 0)
    threshold = 0.05 / 7 / 16 / 24 * (evac_t / 60)
    print(f"\n  Evac {evac_t} min  (threshold = {threshold:.4e} Sv):")
    print(f"    Risk cells = {rc}/{GRID_SIZE**2}")
    if rc > 0:
        print(f"    CVaR max   = {np.max(rm):.4e} Sv")
        print(f"    CVaR mean  = {np.mean(rm[rm>0]):.4e} Sv")

print("\n" + "=" * 65)
print("  All processing complete!")
print("=" * 65)
